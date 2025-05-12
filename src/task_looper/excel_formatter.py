from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelFormatter:

    def __init__(self, outer: "TaskLooperRSCH"):
        self.master_df = outer.df_creator.master_df
        self.df_components = outer.df_creator.df_components

    def run(self, excel_path):
        wb = load_workbook(excel_path)
        self.ws = wb.active

        titles_row_indices = self.get_titles_row_indices()
        orientation_row_indices = self.get_orientation_row_indices()
        coil_row_indices = sorted(self.get_coil_row_indices())
        blank_row_indices = self.get_blank_row_indices()

        # format task title strings
        for idx in titles_row_indices:
            cells_to_format = self.get_cells_in_range(
                idx, idx, 1, self.master_df.shape[1]
            )
            for cell in cells_to_format:
                setattr(cell, "font", Font(bold=True, underline="single", size=15))
                setattr(cell, "alignment", Alignment(horizontal="center"))
            self.ws.merge_cells(self.get_str_range_from_cells(cells_to_format))

        # format coil strings
        for idx in coil_row_indices:
            cells_to_format = self.get_cells_in_range(idx, idx, 1, 1)
            for cell in cells_to_format:
                setattr(cell, "font", Font(bold=True))

        # horizontal border formatting
        for idx in orientation_row_indices:
            cells_to_format = self.get_cells_in_range(
                idx, idx, 1, self.master_df.shape[1]
            )
            for cell in cells_to_format:
                setattr(cell, "border", Border(bottom=Side(style="thick")))

        # vertical border formatting
        cells_to_format = self.get_cells_in_range(1, self.ws.max_row, 1, 1)
        cells_to_format = [
            cell
            for cell in cells_to_format
            if cell.row not in blank_row_indices + coil_row_indices + titles_row_indices
        ]
        for cell in cells_to_format:
            setattr(
                cell,
                "border",
                self.merge_borders(cell.border, Border(right=Side(style="thick"))),
            )

        # set column A width
        self.ws.column_dimensions["A"].width = 20
        wb.save(excel_path)

    def get_titles_row_indices(self):
        title_row_idxs = []
        for target_row in self.df_components["title_rows"]:
            target_row_series = target_row.iloc[0]
            for idx, row in self.master_df.iterrows():
                if row.equals(target_row_series):
                    title_row_idxs.append(idx + 1)  # adjust for 1-indexing of openpyxl
        return title_row_idxs

    def get_orientation_row_indices(self):
        or_row_idxs = []
        target_row_series = self.df_components["orientations_row"].iloc[0]
        for idx, row in self.master_df.iterrows():
            if row.equals(target_row_series):
                or_row_idxs.append(idx + 1)  # adjust for 1-indexing of openpyxl
        return or_row_idxs

    def get_coil_row_indices(self):
        coil_row_idxs = []
        for target_row in self.df_components["coil_rows"].values():
            target_row_series = target_row.iloc[0]
            for idx, row in self.master_df.iterrows():
                if row.equals(target_row_series):
                    coil_row_idxs.append(idx + 1)  # adjust for 1-indexing of openpyxl
        return coil_row_idxs

    def get_blank_row_indices(self):
        blank_row_idxs = []
        target_row_series = self.df_components["blank_row"].iloc[0].astype(object)
        for idx, row in self.master_df.iterrows():
            if row.equals(target_row_series):
                blank_row_idxs.append(idx + 1)
        return blank_row_idxs

    @staticmethod
    def get_str_range_from_cells(cells: list[Cell]):
        min_col = min(cell.column for cell in cells)
        max_col = max(cell.column for cell in cells)
        min_row = min(cell.row for cell in cells)
        max_row = max(cell.row for cell in cells)

        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    def get_cells_in_range(self, start_row, end_row, start_col, end_col):
        cells = [
            cell
            for row in self.ws.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_col,
                max_col=end_col,
            )
            for cell in row
        ]
        return cells

    @staticmethod
    def merge_borders(existing_border: Border, new_border: Border):
        left = new_border.left if new_border.left is not None else existing_border.left
        right = (
            new_border.right if new_border.right is not None else existing_border.right
        )
        top = new_border.top if new_border.top is not None else existing_border.top
        bottom = (
            new_border.bottom
            if new_border.bottom is not None
            else existing_border.bottom
        )
        return Border(left=left, right=right, top=top, bottom=bottom)
