from pathlib import Path
import pandas as pd

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.shared.queueing import get_queue


class DataFrameFormatter:

    def __init__(self, df: pd.DataFrame, excel_path: Path):
        self.df = df
        self.excel_path = excel_path

    def run(self):
        workbook = load_workbook(self.excel_path)
        self.worksheet = workbook.active

        # Add dataframe formatting here

        get_queue().put(("TASK_COMPLETE", "DATAFRAME_FORMATTED"))
