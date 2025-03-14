import os
import sys
sys.path.insert(
    0,
    "C:\\Users\\mliljeroth\\Desktop\\AutoQCtesting\\Scottish-Medium-ACR-Analysis-Framework",
)
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tkinter import Tk, filedialog
from typing import Union
import re
import pandas as pd
import numpy as np

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from platformdirs import user_cache_dir
import pickle

from hazenlib.utils import get_dicom_files, sortDICOMs
from hazenlib.tasks.acr_slice_thickness import ACRSliceThickness
from hazenlib.tasks.acr_uniformity import ACRUniformity
from hazenlib.tasks.acr_snr import ACRSNR
from hazenlib.tasks.acr_geometric_accuracy import ACRGeometricAccuracy
from hazenlib.tasks.acr_spatial_resolution_rsch_test import ACRSpatialResolution

class Mappings:
    IMPLEMENTED_TASKS = (
        "slice_thickness",
        "snr",
        "geometric_accuracy",
        "uniformity",
        "spatial_resolution",
    )

    TASK_CLASSES = (
        ACRSliceThickness,
        ACRSNR,
        ACRGeometricAccuracy,
        ACRUniformity,
        ACRSpatialResolution,
    )

    CLASS_AS_STR = (classx.__name__ for classx in TASK_CLASSES)
    TASK_TO_CLASS = dict(zip(IMPLEMENTED_TASKS, TASK_CLASSES))
    CLASS_STR_TO_TASK = dict(zip(CLASS_AS_STR, IMPLEMENTED_TASKS))


class StrUtils:

    @staticmethod
    def print_break(length: int):
        print("".join("-" for _ in range(length)))

    @staticmethod
    def extract_coil(file_path: str):
        if isinstance(file_path, list):
            file_path = file_path[0]
        file_path = os.path.basename(file_path)
        return file_path[: file_path.find("_")]

    @staticmethod
    def extract_orientation(file_path: str):
        if isinstance(file_path, list):
            file_path = file_path[0]
        file_path = os.path.basename(file_path)
        underscore_locs = [match.start() for match in re.finditer(re.escape("_"), file_path)]
        slice_start = underscore_locs[0] + 1
        slice_end = None if len(underscore_locs) == 1 else underscore_locs[1]
        return file_path[slice_start:slice_end]


class CacheManager:

    def __init__(self, app_name: str):
        self.cache_dir = user_cache_dir(app_name)
        os.makedirs(self.cache_dir, exist_ok=True)

    def store_cache(self, data, cache_name: str):
        cache_file = os.path.join(self.cache_dir, cache_name)
        try: 
            with open(cache_file, "wb") as file:
                pickle.dump(data, file)
                print(f"Data successfully stored to cache: {cache_file}")
                
        except Exception as e:
            raise IOError(f"Error storing to cache: {e}")

    def load_cache(self, cache_name: str):
        cache_file = os.path.join(self.cache_dir, cache_name)
        try:
            with open(cache_file, "rb") as file:
                data = pickle.load(file)
                print(f"Pulling data from cache from cache name '{cache_name}'!")
                return data
            
        except (pickle.UnpicklingError, EOFError) as e:
            raise IOError(f"Error loading cache: {e}")

class TaskLooperRSCH:
    APP_NAME = "AutoQC_ACR"

    def run(self, tasks: list, pull_from_cache: bool = False):
        self.check_for_unsupported_tasks(tasks)
        self.cache_manager = CacheManager(self.APP_NAME)

        # handling IO folder logic
        StrUtils.print_break(20)
        if pull_from_cache:
            print("Attempting to pull previous results from cache.")
            self.folder_io = self.cache_manager.load_cache("folder_io")
        else:
            print("Asking user to select top-level I/O folders.")
            self.folder_io = self.FolderIO()
            self.folder_io.run()
            StrUtils.print_break(20)
        self.all_keys = self.generate_keys_dict()

        # Handling results generation logic
        if pull_from_cache:
            self.results = self.cache_manager.load_cache("results")
            StrUtils.print_break(20)
        else:
            print("Running Hazen tasks.")
            self.task_runner = self.TaskRunner(outer=self)
            self.results = self.task_runner.run()
            StrUtils.print_break(20)
            print("Storing results in cache.")
            self.cache_manager.store_cache(self.folder_io, "folder_io")
            self.cache_manager.store_cache(self.results, "results")
            StrUtils.print_break(20)

        # handling results creation logic
        self.excel_path = os.path.join(self.folder_io.out_top, "output_AutoQC_ACR.xlsx")
        self.df_creator = self.DataFrameCreator(outer=self)
        self.df_creator.run(self.excel_path)

        self.excel_formatter = self.ExcelFormatter(outer=self)
        self.excel_formatter.run(self.excel_path)
        print(f"Results successfully stored at: {self.excel_path}")
        StrUtils.print_break(20)


    def check_for_unsupported_tasks(self, tasks):
        key_check = [task in Mappings.TASK_TO_CLASS for task in tasks]
        if all(key_check):
            self.tasks = tasks
        else:
            err = str(
                f"Task {tasks[key_check.index(False)]} not implemented! Available tasks: "
                f"{', '.join(list(Mappings.TASK_TO_CLASS.keys()))}"
            )
            raise KeyError(err)

    def generate_keys_dict(self) -> dict:
        task_keys = [f"{task}" for task in self.tasks]
        coil_keys = list(set([StrUtils.extract_coil(d) for d in self.folder_io.out_dirs]))
        orientation_keys = list(set([StrUtils.extract_orientation(d) for d in self.folder_io.in_dirs]))
        all_keys = {"tasks": task_keys, "coils": coil_keys, "orientations": orientation_keys}
        return all_keys

    class FolderIO:

        FOLDERS_TO_EXCLUDE = ["_Excluded Images", "_ND Images"]

        def run(self):
            self.in_top = self.get_in_top_dir()
            self.out_top = self.get_out_top_dir()
            self.in_dirs = self.sort_dicoms()
            self.out_dirs = self.gen_out_dirs()

        @staticmethod
        def get_in_top_dir():
            root = Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            in_top = filedialog.askdirectory(
                parent=root, title="Choose top level input data folder."
            )
            if in_top == "":
                root.destroy()
                raise ValueError("No top level input folder selected.")
            if len(os.listdir(in_top)) == 0:
                root.destroy()
                raise ValueError(
                    "The selected top level input folder should not be completely empty."
                )
            return in_top

        @staticmethod
        def get_out_top_dir():
            root = Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            out_top = filedialog.askdirectory(
                parent=root, title="Choose top level output data folder."
            )
            if out_top == "":
                root.destroy()
                raise ValueError("No top level output folder selected.")
            root.destroy()
            return out_top

        def sort_dicoms(self):
            in_dirs = [
                item
                for item in os.listdir(self.in_top)
                if os.path.isdir(os.path.join(self.in_top, item))
            ]
            if len(in_dirs) == 0:
                sortDICOMs(self.in_top)
                in_dirs = [
                    item
                    for item in os.listdir(self.in_top)
                    if os.path.isdir(os.path.join(self.in_top, item))
                ]
            elif len(in_dirs) == 9 + len(self.FOLDERS_TO_EXCLUDE):
                pass
            else:
                raise ValueError(
                    "There should only be DICOM images in the top level input folder when running this programme, \n or the images should already be sorted."
                )
            in_dirs = [
                os.path.join(self.in_top, x)
                for x in in_dirs
                if not any([x in folder for folder in self.FOLDERS_TO_EXCLUDE])
            ]
            return in_dirs

        def gen_out_dirs(self):
            out_dirs = []
            for in_dir in self.in_dirs:
                if not any(x in in_dir for x in self.FOLDERS_TO_EXCLUDE):
                    out_dir = f"{self.out_top}/{os.path.basename(in_dir)}Results"
                    if os.path.exists(out_dir):
                        print(f"Overwriting results folder: {out_dir}")
                    else:
                        os.makedirs(out_dir)
                        print(f"Generating results fodler: {out_dir}")
                    out_dirs.append(out_dir)
            return out_dirs

    class TaskRunner:
        def __init__(self, outer: "TaskLooperRSCH"):
            self.outer = outer

        def run(self):
            self.results = self.init_results_dict()
            self.populate_results_dict()
            return self.results

        def init_results_dict(self) -> dict:
            results = {
                tk: {
                    ck: {orientation: [] for orientation in self.outer.all_keys["orientations"]}
                    for ck in self.outer.all_keys["coils"]
                }
                for tk in self.outer.all_keys["tasks"]
            }
            return results

        def populate_results_dict(self):
            # get raw results output from hazen code
            raw_results = [
                self._run_tasks_on_folder(in_dir, out_dir, tasks=self.outer.tasks)
                for (in_dir, out_dir) in zip(self.outer.folder_io.in_dirs, self.outer.folder_io.out_dirs)
            ]

            # populate results dict (organised results)
            for outer_list in raw_results:
                for inner_dict in outer_list:
                    task_key = Mappings.CLASS_STR_TO_TASK[inner_dict["task"]]
                    coil_key = StrUtils.extract_coil(inner_dict["file"])
                    orientation_key = StrUtils.extract_orientation(inner_dict["file"])
                    self.results[task_key][coil_key][orientation_key] = inner_dict

        @staticmethod
        def _run_tasks_on_folder(in_dir: str, out_dir: str, tasks: list) -> list[dict]:
            input_data = get_dicom_files(os.path.join(in_dir))
            task_objs = [
                Mappings.TASK_TO_CLASS[task](
                    input_data=input_data, report_dir=out_dir, report=True, MediumACRPhantom=True
                )
                for task in tasks
            ]
            results_folder = [taskobj.run() for taskobj in task_objs]
            return results_folder

    class DataFrameCreator:

        def __init__(self, outer: "TaskLooperRSCH"):
            self.outer = outer
            self.df_width = len(self.outer.all_keys["orientations"]) + 1
            self.df_components = {
                "title_rows": [
                    pd.DataFrame([task.upper()] + [np.nan for x in range(self.df_width - 1)]).T
                    for task in self.outer.tasks
                ],
                "blank_row": pd.DataFrame(
                    np.nan, index=range(1), columns=range(self.df_width)
                ),
                "orientations_row": pd.DataFrame([np.nan] + self.outer.all_keys["orientations"]).T,
                "coil_rows": {
                    coil: pd.DataFrame([coil] + [np.nan] * (self.df_width - 1)).T
                    for coil in self.outer.all_keys["coils"]
                }
            }

        def run(self, excel_path: str) -> pd.DataFrame:
            task_dataframes = {
                task: pd.concat(
                    self._list_with_delimiter(
                        lst=[
                            getattr(self, f"_populate_{task}")(
                                modular_df_coil=self._construct_modular_df_for_coil(coil=coil),
                                data=coil_dict,
                            )
                            for coil, coil_dict in task_dict.items()
                        ],
                        delim=self.df_components["blank_row"],
                    ),
                    ignore_index=True,
                )
                for task, task_dict in self.outer.results.items()
            }

            zipped = zip(
                self.df_components["title_rows"],
                list(task_dataframes.values()),
                (self.df_components["blank_row"] for task in self.outer.tasks),
            )

            self.master_df = pd.concat(
                [elem for triplet in zipped for elem in triplet], ignore_index=True
            )
            self.master_df.to_excel(excel_path, header=False, index=False, sheet_name="Sheet1")

        def _construct_modular_df_for_coil(self, coil: str):
            modular_df_coil = pd.concat(
                [self.df_components["coil_rows"][coil], self.df_components["orientations_row"]],
                ignore_index=True,
            )
            return modular_df_coil

        @staticmethod
        def _populate_slice_thickness(modular_df_coil: pd.DataFrame, data: dict):
            slice_thicknesses = [
                data[orientation].get("measurement", {}).get("slice width mm", pd.NA)
                for orientation in modular_df_coil.iloc[1, 1:]
            ]
            perc_diff_to_set = [(st - 5) / 5 * 100 for st in slice_thicknesses]

            to_add = pd.DataFrame(
                [
                    ["Slice Thickness (mm)"] + slice_thicknesses,
                    ["% Diff to set (5mm)"] + perc_diff_to_set,
                ]
            )

            return pd.concat([modular_df_coil, to_add], ignore_index=True)

        @staticmethod
        def _populate_snr(modular_df_coil: pd.DataFrame, data: dict):
            measured_and_normalised_snrs = [
                (
                    data[orientation]
                    .get("measurement", {})
                    .get("snr by smoothing", {})
                    .get("measured", pd.NA),
                    data[orientation]
                    .get("measurement", {})
                    .get("snr by smoothing", {})
                    .get("normalised", pd.NA),
                )
                for orientation in modular_df_coil.iloc[1, 1:]
            ]
            measured_snrs, normalised_snrs = zip(*measured_and_normalised_snrs)

            to_add = pd.DataFrame(
                [["Image SNR"] + list(measured_snrs), ["Normalised SNR"] + list(normalised_snrs)]
            )

            return pd.concat([modular_df_coil, to_add], ignore_index=True)

        @staticmethod
        def _populate_geometric_accuracy(modular_df_coil: pd.DataFrame, data: dict):
            true_length = 173

            def get_perc_diff_and_cv(lengths):
                lengths = list(lengths)
                perc_differences = [(1 - length / true_length) * 100 for length in lengths]
                av_perc_diff = np.mean(perc_differences)
                cv = np.std(lengths) / np.mean(lengths) * 100
                return av_perc_diff, cv

            perc_diffs_and_cvs = [
                get_perc_diff_and_cv(data[orientation].get("measurement").values())
                for orientation in modular_df_coil.iloc[1, 1:]
            ]
            perc_diffs, cvs = zip(*perc_diffs_and_cvs)

            to_add = pd.DataFrame(
                [["% Distortion"] + list(cvs), ["% Diff to actual"] + list(perc_diffs)]
            )

            return pd.concat([modular_df_coil, to_add], ignore_index=True)

        @staticmethod
        def _populate_uniformity(modular_df_coil: pd.DataFrame, data: dict):
            uniformities = [
                data[orientation].get("measurement", {}).get("integral uniformity %", pd.NA)
                for orientation in modular_df_coil.iloc[1, 1:]
            ]

            to_add = pd.DataFrame(["% Integral Uniformity"] + uniformities).T

            return pd.concat([modular_df_coil, to_add], ignore_index=True)

        @staticmethod
        def _populate_spatial_resolution(modular_df_coil: pd.DataFrame, data: dict):
            raise ValueError("Spatial resolution not yet implemented!")

        @staticmethod
        def _list_with_delimiter(lst: list, delim: any):
            delimited_list = [
                elem for pair in zip(lst, [delim] * (len(lst) - 1)) for elem in pair
            ] + [lst[-1]]

            return delimited_list

    class ExcelFormatter:

        def __init__(self, outer: "TaskLooperRSCH"):
            self.master_df = outer.df_creator.master_df
            self.df_components = outer.df_creator.df_components

        def run(self, excel_path):
            wb = load_workbook(excel_path)
            self.ws = wb.active

            titles_row_indices = self.get_titles_row_indices()
            orientation_row_indices = self.get_orientation_row_indices()
            coil_row_indices = sorted(self.get_coil_row_indices())
            blank_row_indices = self.get_blank_row_indices()

            # format task title strings
            for idx in titles_row_indices:
                cells_to_format = self.get_cells_in_range(idx, idx, 1, self.master_df.shape[1])
                for cell in cells_to_format:
                    setattr(cell, "font", Font(bold=True, underline="single", size=15))
                    setattr(cell, "alignment", Alignment(horizontal="center"))
                self.ws.merge_cells(self.get_str_range_from_cells(cells_to_format))

            # format coil strings
            for idx in coil_row_indices:
                cells_to_format = self.get_cells_in_range(idx, idx, 1, 1)
                for cell in cells_to_format:
                    setattr(cell, "font", Font(bold=True))

            # horizontal border formatting
            for idx in orientation_row_indices:
                cells_to_format = self.get_cells_in_range(idx, idx, 1, self.master_df.shape[1])
                for cell in cells_to_format:
                    setattr(cell, "border", Border(bottom=Side(style="thick")))

            # vertical border formatting
            cells_to_format = self.get_cells_in_range(1, self.ws.max_row, 1, 1)
            cells_to_format = [cell for cell in cells_to_format if cell.row not in blank_row_indices + coil_row_indices + titles_row_indices]
            for cell in cells_to_format:
                setattr(cell, "border", self.merge_borders(cell.border, Border(right=Side(style="thick"))))

            # set column A width
            self.ws.column_dimensions["A"].width = 20
            wb.save(excel_path)

        def get_titles_row_indices(self):
            title_row_idxs = []
            for target_row in self.df_components["title_rows"]:
                target_row_series = target_row.iloc[0]
                for idx, row in self.master_df.iterrows():
                    if row.equals(target_row_series):
                        title_row_idxs.append(idx + 1) # adjust for 1-indexing of openpyxl
            return title_row_idxs

        def get_orientation_row_indices(self):
            or_row_idxs = []
            target_row_series = self.df_components["orientations_row"].iloc[0]
            for idx, row in self.master_df.iterrows():
                if row.equals(target_row_series):
                    or_row_idxs.append(idx + 1) # adjust for 1-indexing of openpyxl
            return or_row_idxs

        def get_coil_row_indices(self):
            coil_row_idxs = []
            for target_row in self.df_components["coil_rows"].values():
                target_row_series = target_row.iloc[0]
                for idx, row in self.master_df.iterrows():
                    if row.equals(target_row_series):
                        coil_row_idxs.append(idx + 1) # adjust for 1-indexing of openpyxl
            return coil_row_idxs
        
        def get_blank_row_indices(self):
            blank_row_idxs = []
            target_row_series = self.df_components["blank_row"].iloc[0].astype(object)
            for idx, row in self.master_df.iterrows():
                if row.equals(target_row_series):
                    blank_row_idxs.append(idx + 1)
            return blank_row_idxs

        @staticmethod
        def get_str_range_from_cells(cells: list[Cell]):
            min_col = min(cell.column for cell in cells)
            max_col = max(cell.column for cell in cells)
            min_row = min(cell.row for cell in cells)
            max_row = max(cell.row for cell in cells)

            return (
                f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            )
        
        def get_cells_in_range(self, start_row, end_row, start_col, end_col):
            cells = [
                cell for row in self.ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col)
                for cell in row 
            ]
            return cells

        @staticmethod
        def merge_borders(existing_border: Border, new_border: Border):
            left = new_border.left if new_border.left is not None else existing_border.left
            right = new_border.right if new_border.right is not None else existing_border.right
            top = new_border.top if new_border.top is not None else existing_border.top
            bottom = new_border.bottom if new_border.bottom is not None else existing_border.bottom
            return Border(left=left, right=right, top=top, bottom=bottom)


task_looper_rsch = TaskLooperRSCH()
task_looper_rsch.run(
    tasks=[
        "slice_thickness",
        "snr",
        "geometric_accuracy",
        "uniformity",
    ],
    pull_from_cache=True
)
